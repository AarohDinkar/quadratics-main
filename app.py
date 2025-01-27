import os
import uuid
import datetime
from functools import wraps
import re
import base64
import bcrypt
import io
import logging
from pathlib import Path
import json

from flask import Flask, request, jsonify, render_template, redirect, url_for, session, flash, make_response
from werkzeug.utils import secure_filename

from google.cloud import vision
from google.oauth2 import service_account
from google.cloud import firestore
from google.cloud import storage

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from bs4 import BeautifulSoup
from google.protobuf.json_format import MessageToDict

from dotenv import load_dotenv

# ---------------- Anthropic (Claude) Import -------------------
import anthropic

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
app.secret_key = os.urandom(24)  # Replace with a secure key in production

# Load environment variables
load_dotenv()

###############################################################################
#                       ENVIRONMENT VARIABLE CHECKS
###############################################################################
required_vars = [
    "VISION_CREDENTIALS",   
    "FIRESTORE_CREDENTIALS",
    "STORAGE_CREDENTIALS",
    "PROJECT_ID",
    "ANTHROPIC_API_KEY",    # Now using Anthropic instead of Gemini
    "BUCKET_NAME"
]
missing_vars = [var for var in required_vars if not os.getenv(var)]
if missing_vars:
    raise EnvironmentError(f"Missing required environment variables: {', '.join(missing_vars)}")

PROJECT_ID = os.getenv("PROJECT_ID")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
BUCKET_NAME = os.getenv("BUCKET_NAME")

###############################################################################
#                    SERVICE ACCOUNT LOADING UTILITY
###############################################################################
def load_service_account_credentials(env_var):
    creds_path = os.getenv(env_var)
    if not creds_path:
        raise ValueError(f"Missing environment variable: {env_var}")
    return service_account.Credentials.from_service_account_file(creds_path)

###############################################################################
#                      INITIALIZE GOOGLE CLOUD CLIENTS
###############################################################################
try:
    # --- Cloud Vision Client ---
    vision_credentials = load_service_account_credentials("VISION_CREDENTIALS")
    vision_client = vision.ImageAnnotatorClient(credentials=vision_credentials)

    # --- Anthropic Client (Claude) ---
    claude_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    # --- Firestore Client ---
    firestore_credentials = load_service_account_credentials("FIRESTORE_CREDENTIALS")
    db = firestore.Client(project=PROJECT_ID, credentials=firestore_credentials)

    # --- Cloud Storage Client ---
    storage_credentials = load_service_account_credentials("STORAGE_CREDENTIALS")
    storage_client = storage.Client(project=PROJECT_ID, credentials=storage_credentials)
    bucket = storage_client.bucket(BUCKET_NAME)

except Exception as e:
    logger.error(f"Failed to initialize clients: {str(e)}")
    raise

###############################################################################
#              HELPER FUNCTIONS AND UTILITY (PASSWORD, ETC.)
###############################################################################
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

def allowed_file(filename):
    """Check if the uploaded file has an allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def hash_password(password: str) -> str:
    """Hash a plaintext password using bcrypt."""
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode('utf-8'), salt)
    return hashed.decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    """Verify a plaintext password against the hashed version."""
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def login_required(f):
    """Decorator to require login for routes."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def clean_text(text: str) -> str:
    """Remove excessive asterisks, quotes and unnecessary formatting from text."""
    text = re.sub(r'\*{2,}', '', text)
    text = re.sub(r'\'{2,}', '', text)
    text = re.sub(r'^\*+|\*+$', '', text, flags=re.MULTILINE)
    text = re.sub(r"^'+|'+$", '', text, flags=re.MULTILINE)
    return text.strip()

###############################################################################
#            FUNCTION TO REMOVE UNWANTED OCR FIELDS FROM VISION OUTPUT
###############################################################################
def remove_unwanted_ocr_data(ocr_dict: dict) -> dict:
    """
    Given the raw OCR dictionary from Cloud Vision, remove unwanted fields:
    top-level keys (cropHintsAnnotation, imagePropertiesAnnotation, safeSearchAnnotation, labelAnnotations),
    confidence/bounding boxes in deeper levels, etc. Returns a cleaned dictionary.
    """
    # Copy so we don't mutate the original
    cleaned_data = json.loads(json.dumps(ocr_dict))  

    # Remove top-level keys
    keys_to_remove_top = ['cropHintsAnnotation', 'imagePropertiesAnnotation', 'safeSearchAnnotation', 'labelAnnotations']
    for key in keys_to_remove_top:
        cleaned_data.pop(key, None)

    # Remove page-level unwanted info
    if ('fullTextAnnotation' in cleaned_data and
        'pages' in cleaned_data['fullTextAnnotation'] and
        len(cleaned_data['fullTextAnnotation']['pages']) > 0):
        page = cleaned_data['fullTextAnnotation']['pages'][0]
        page.pop('confidence', None)
        page.pop('property', None)

        # Remove block/paragraph/word-level boundingBoxes & confidences
        if 'blocks' in page:
            for block in page['blocks']:
                block.pop('blockType', None)
                block.pop('confidence', None)
                block.pop('boundingBox', None)
                block.pop('property', None)
                if 'paragraphs' in block:
                    for paragraph in block['paragraphs']:
                        paragraph.pop('boundingBox', None)
                        paragraph.pop('confidence', None)
                        if 'words' in paragraph:
                            for word in paragraph['words']:
                                word.pop('boundingBox', None)
                                word.pop('confidence', None)
                                word.pop('property', None)
                                # remove symbols entirely
                                if 'symbols' in word:
                                    word.pop('symbols', None)

    return cleaned_data if isinstance(cleaned_data, dict) else {} # Ensure that the function returns a dict, or an empty dict to avoid issues later on
###############################################################################
#             FUNCTION TO CALL VISION, RETURN RAW & CLEANED OCR
###############################################################################
def process_image_with_vision(file_bytes: bytes) -> dict:
    """
    Calls Cloud Vision API (DOCUMENT_TEXT_DETECTION) on the given image bytes,
    returns a dict containing:
      - 'ocr_data': the raw OCR dictionary (Vision response),
      - 'cleaned_ocr_data': the sanitized version of that OCR dictionary.
    """
    try:
        image = vision.Image(content=file_bytes)
        response = vision_client.document_text_detection(image=image)
        # Convert protobuf response to a Python dict
        response_dict = MessageToDict(response._pb)
        
        # Now clean up the response to remove bounding boxes, confidences, etc.
        cleaned_data = remove_unwanted_ocr_data(response_dict)

        return {
            "ocr_data": response_dict,
            "cleaned_ocr_data": cleaned_data
        }
    except Exception as e:
        logger.error(f"Error processing image with Cloud Vision: {str(e)}")
        raise

###############################################################################
#           CLAUDE CALL - USE CLEANED OCR DATA FOR STRUCTURED ANALYSIS
###############################################################################
def generate_structured_analysis(cleaned_ocr_data: dict, file_bytes: bytes):
    """
    Use Anthropic’s Claude model to generate a structured analysis in JSON format.
    We pass the cleaned_ocr_data into the prompt.
    """
    try:
        system_prompt = """
You are an expert in analyzing electrical panel diagrams. Your task is to extract information from circuit diagrams and structure it into an JSON format for easy visualization. You will be given OCR data of circuit diagrams. The OCR data includes boundingPoly, vertices, and description, which represent the position and content of each text element in the diagram. 

Your primary goal is to understand the structure of an electrical panel based on the components and connections shown.

Your main task is to go from top to bottom and identify the components and their relationships.

**Instructions:**

1.  Diagram Analysis: Use the OCR data (which includes boundingPoly, vertices, and description) to identify components. Prioritize the understanding of component types and relationships from the image rather than strictly following a top-down approach. Use spatial relationships where available but don't rely just on the spatial positions.

2.  Component Identification and Relationships:
    *  For each component, extract the component's name, quantity, and specifications from the OCR data.
    *  Use the component names and associated specifications, their text content, along with the visual connections of lines in the images, to establish relationships.
    * For each component, combine the component's name and specifications into a single column called "name" in the JSON output.
    * Identify components based on proximity as well. Sometimes components are written in two lines vertically but are part of the same component. In such cases, combine the text from both lines to form the complete component name and specifications.

3.  How to use the OCR data:
    * OCR data has 4 paramters : 1)description  2) bounding_boxes 3) component_count 4) detailed_components
    * description: The text extracted from the image (e.g., "MCB," "630A, 4P").
    * boundingPoly: The bounding box (rectangle) around the extracted text, showing its position on the image.
    * vertices: The coordinates (x, y) of the four corners of the bounding box, used to determine the text's position (top, bottom, left, right) on the diagram.

4. Component Identification and Relationships:
   * For components written vertically, use these rules to combine text:
     - If text elements are vertically stacked (aligned on x-axis within 10 pixels)
     - If vertical spacing between elements is less than 30 pixels
     - If text elements represent specifications of the same component
   * Examples of text to combine:
     "63A, FP" + "MCB" + "10kA" → "63A, FP MCB 10kA"
     "100A" + "MCCB" → "100A MCCB"
   * Common patterns to look for:
     - Rating + Type (e.g., "63A" + "MCB")
     - Type + Specifications (e.g., "MCB" + "10kA")
     - Rating + Poles + Type (e.g., "63A, FP" + "MCB")
   *  Use the OCR data's co-ordinates to carefully combine text elements that belong to the same component.  

5. OCR Data Processing Rules:
   * Consider text elements as part of the same component if:
     - They share vertical alignment (x-coordinate ±10 pixels)
     - Vertical gap is less than 30 pixels
     - Content follows standard component patterns
   * Always combine specifications with their base component
   * Order specifications in standard format: Rating > Poles > Type > Details    

6. How to measure the quantity:
  * The quantity of each component can be determined by counting the number of times the component's name appears in the OCR data.
  * Ensure that the component names are matched accurately to avoid counting errors.
  * If the quantity cannot be determined from the OCR data, mark the quantity as "Unknown" in the JSON output.
  *
  6.1 Utilizing "text" keys in OCR data:
    * The "text" keys under fullTextAnnotation (fullTextAnnotation.text) contain the extracted text from the image.
    * Use these "text" keys to identify and count the occurrences of each component's name.
    * Ensure that the text is parsed accurately to avoid counting errors.
    * If the text is split across multiple lines, combine the lines to form the complete component name and specifications.

7. How to identify the title correctly:
  * Titles are mostly located at the top left of the diagram, often inside a small box.
  * Ensure that the title is included in the JSON output rather than placing it in the additional notes section.
  * Use the boundingPoly and vertices to accurately identify and extract the title from the OCR data.

8.  **Additional Notes:** Include any relevant information that was not captured in the rows above, in a list of string format. If no other relevant information to include, then include `": Everything looking good."` as a single string in the array. Make sure, the additional notes are included within the JSON output. Use simple bullet points for clarity for each point. For instance use - before every point.

9.  **Output:** Return the data in the format of the JSON, where all the data, including the Additional Notes is inside the JSON output.



10. **JSON Format:** The JSON format should be as follows:

{
  "title": "OCR based title",
  "rows": [
    {
      "serial_number": 1,
      "name": "MCCB 150/4P",
      "quantity": "10"
    },
    ...
  ],
  "additional_notes": [
    "- Example",
    ...
  ]
}


8.1 **Additional Notes:
 -unclear components: If any component name or type cannot be identified clearly, include the component name and the issue in this section.
 -unclear specifications: If any specification or detail cannot be identified clearly, include the specification, the related component name, and the issue in this section.
 -quantity issues: If there are any discrepancies in the expected and found quantities of components, include the component name, the expected quantity, the found quantity, and the issue in this section
 -recognition issues: If there are any issues with the OCR text recognition, include the raw text and a possible interpretation in this section.   

In additional_notes : If any component, specification, or quantity cannot be clearly identified and categorized into one of the three sections, include the component or the un-parsed data in a separate section called "Additional Notes" below the JSON output in the text format. Use simple bullet points for clarity. For instance use - before every point. If no component needs to be included in additional notes, don't include it. But mention ": Everything looking good.", if you don't have any instruction or anything to be include here.

Be sure your final answer is valid JSON, parseable by Python. **No extra text** outside the JSON.

Make the output clean, hierarchical, and easy to understand. Don’t leave anything, consider everything!
        """.strip()

        user_content = (
            f"Here is the OCR data (JSON):\n\n"
            f"{json.dumps(cleaned_ocr_data, indent=2)}"
        )

        message = claude_client.messages.create(
            model="claude-3-5-sonnet-20241022",
            system=system_prompt,
            max_tokens=8192,
            messages=[
                {"role": "user", "content": user_content}
            ]
        )

        raw_response = message.content

        # If we get a list of partial TextBlock objects, convert them to strings
        if isinstance(raw_response, list):
            new_lines = []
            for block in raw_response:
                # Attempt to extract text from TextBlock objects
                if hasattr(block, "content") and isinstance(block.content, str):
                    new_lines.append(block.content)
                elif hasattr(block, "text") and isinstance(block.text, str):
                    new_lines.append(block.text)
                else:
                    # Fallback to string conversion
                    new_lines.append(str(block))
            # Join them into one string
            raw_response = "\n".join(new_lines)
        elif not isinstance(raw_response, str):
            # If raw_response is not a string or list, convert it to string
            raw_response = str(raw_response)

        # Now apply your existing cleaning logic
        cleaned_response = clean_text(raw_response)
        cleaned_response = re.sub(r'^```(?:json|.+)?\n?', '', cleaned_response)
        cleaned_response = re.sub(r'\n?```$', '', cleaned_response)

        return cleaned_response

    except Exception as e:
        logger.error(f"Error generating structured analysis with Claude: {str(e)}")
        raise


###############################################################################
#                            AUTH & MAIN ROUTES
###############################################################################
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Render and handle the login page."""
    if request.method == 'POST':
        username = request.form.get('username').strip()
        password = request.form.get('password').strip()

        if not username or not password:
            flash('Please enter both username and password.', 'error')
            return redirect(url_for('login'))

        try:
            user_ref = db.collection('users').document(username)
            user_doc = user_ref.get()
            if not user_doc.exists or not verify_password(password, user_doc.to_dict().get('password')):
                flash('Invalid username or password.', 'error')
                return redirect(url_for('login'))
            session['username'] = username
            flash('Logged in successfully!', 'login_success')
            return redirect(url_for('index'))

        except Exception as e:
            logger.error(f"Login error: {str(e)}")
            flash('An error occurred during login. Please try again.', 'error')
            return redirect(url_for('login'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Log the user out."""
    session.pop('username', None)
    flash('Logged out successfully.', 'logout_success')
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    """Render the main page for authenticated users,
    and retrieve the last modified thread (if any)."""
    username = session['username']
    threads_query = (
        db.collection('users').document(username).collection('threads')
        .order_by('created_at', direction=firestore.Query.DESCENDING).limit(1).stream()
    )
    last_thread = next((doc.to_dict().get('thread_name') for doc in threads_query), None)
    return render_template('index.html', username=username, last_thread=last_thread)

###############################################################################
#                          THREADS & IMAGES ROUTES
###############################################################################
@app.route('/getThreads', methods=['GET'])
@login_required
def get_threads():
    """Returns a list of threads for the current user."""
    try:
        username = session['username']
        threads_ref = (
            db.collection('users').document(username).collection('threads')
              .order_by('created_at', direction=firestore.Query.DESCENDING).stream()
        )
        threads = [
            {
                'thread_name': doc.to_dict().get('thread_name'),
                'created_at': doc.to_dict().get('created_at').isoformat() if doc.to_dict().get('created_at') else None
            }
            for doc in threads_ref
        ]
        return jsonify(threads)
    except Exception as e:
        logger.error(f"Error retrieving threads: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/getThreadImages', methods=['GET'])
@login_required
def get_thread_images():
    """Returns a list of images (with their outputs) for a given threadName."""
    try:
        username = session['username']
        thread_name = request.args.get('threadName')
        if not thread_name:
            return jsonify({'error': 'threadName query parameter is required'}), 400

        thread_ref = db.collection('users').document(username).collection('threads').document(thread_name)
        thread_doc = thread_ref.get()
        if not thread_doc.exists:
            return jsonify({'error': f'Thread "{thread_name}" not found'}), 404

        images_collection = thread_ref.collection('images').stream()
        images_list = []
        for img_doc in images_collection:
            img_data = img_doc.to_dict()
            gcs_path = img_data.get('gcs_path')
            image_url = ''
            if gcs_path and gcs_path.startswith('gs://'):
                parts = gcs_path.replace('gs://', '').split('/', 1)
                if len(parts) == 2:
                    bucket_name, file_name_gcs = parts
                    image_url = f"https://storage.googleapis.com/{bucket_name}/{file_name_gcs}"

            # Fetch latest structured analysis
            latest_analysis = img_data.get('structured_analysis', '')
            latest_index = -1
            for key in img_data:
                match = re.match(r'structured_analysis_(\d+)', key)
                if match:
                    idx = int(match.group(1))
                    if idx > latest_index:
                        latest_index = idx
                        latest_analysis = img_data.get(key)

            images_list.append({
                'file_name': img_data.get('file_name'),
                'ocr_data': img_data.get('ocr_data'),
                'cleaned_ocr_data': img_data.get('cleaned_ocr_data'),
                'structured_analysis': img_data.get('structured_analysis'),
                'structured_analysis_json': img_data.get('structured_analysis_json'),
                'latest_structured_analysis': latest_analysis,
                'gcs_path': gcs_path,
                'image_url': image_url
            })

        return jsonify(images_list), 200

    except Exception as e:
        logger.error(f"Error retrieving thread images: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/updateThreadName', methods=['POST'])
@login_required
def update_thread_name():
    """Update the name of an existing thread."""
    try:
        username = session['username']
        data = request.get_json()
        current_name = data.get('current_name')
        new_name = data.get('new_name')

        if not current_name or not new_name:
            return jsonify({'error': 'Both current_name and new_name are required.'}), 400

        existing_thread_ref = db.collection('users').document(username).collection('threads').document(new_name)
        if existing_thread_ref.get().exists:
            return jsonify({'error': 'The new thread name already exists.'}), 400

        threads_stream = (
            db.collection('users').document(username).collection('threads')
            .where('thread_name', '==', current_name).stream()
        )
        threads = list(threads_stream)

        if not threads:
            return jsonify({'error': 'Thread not found.'}), 404

        thread_doc = threads[0].reference
        thread_doc.update({'thread_name': new_name})
        logger.info(f"Thread name updated from '{current_name}' to '{new_name}' for user '{username}'.")
        return jsonify({'message': 'Thread name updated successfully.'}), 200

    except Exception as e:
        logger.error(f"Error updating thread name: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/checkThreadName', methods=['GET'])
@login_required
def check_thread_name():
    """Check if a thread name already exists."""
    try:
        username = session['username']
        thread_name = request.args.get('name', '').strip()
        if not thread_name:
            return jsonify({'exists': False}), 200
        thread_ref = db.collection('users').document(username).collection('threads').document(thread_name)
        return jsonify({'exists': thread_ref.get().exists}), 200

    except Exception as e:
        logger.error(f"Error checking thread name: {str(e)}")
        return jsonify({'error': str(e)}), 500

###############################################################################
#                     UPLOAD & PROCESS (VISION -> CLAUDE)
###############################################################################
@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    """Handle file upload, Cloud Vision, Claude, Firestore save."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    thread_name = request.form.get('thread_name')

    if not thread_name or not thread_name.strip():
        return jsonify({'error': 'Thread name is required and cannot be empty.'}), 400

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        username = session['username']
        blob = bucket.blob(f"{username}/{thread_name}/{filename}")
        blob.upload_from_file(file, content_type=file.content_type)
        gcs_path = f"gs://{BUCKET_NAME}/{username}/{thread_name}/{filename}"

        # Make sure to re-seek the file to read bytes again
        file.seek(0)
        file_bytes = file.read()
        if not file_bytes:
            # If file buffer is empty, download from GCS (rare edge case)
            file_bytes = blob.download_as_bytes()

        try:
            # Get raw & cleaned OCR data from Vision
            ocr_results = process_image_with_vision(file_bytes)
            raw_ocr_data = ocr_results["ocr_data"]
            cleaned_ocr_data = ocr_results["cleaned_ocr_data"]

            # Generate structured analysis using Claude on cleaned_ocr_data
            structured_analysis = generate_structured_analysis(cleaned_ocr_data, file_bytes)

            # Upsert thread doc in Firestore
            thread_ref = db.collection('users').document(username).collection('threads').document(thread_name)
            thread_ref.set({'thread_name': thread_name, 'created_at': firestore.SERVER_TIMESTAMP}, merge=True)

            # Add a new image doc within this thread
            image_id = str(uuid.uuid4())
            images_ref = thread_ref.collection('images').document(image_id)
            images_ref.set({
                'file_name': filename,
                'gcs_path': gcs_path,
                'ocr_data': raw_ocr_data,          
                'cleaned_ocr_data': cleaned_ocr_data,
                'structured_analysis': structured_analysis,
                'uploaded_at': datetime.datetime.utcnow()
            })

            return jsonify({'structured_analysis': structured_analysis})

        except Exception as e:
            logger.error(f"Upload error: {str(e)}")
            return jsonify({'error': str(e)}), 500
    else:
        return jsonify({'error': 'Invalid file format. Only PNG, JPG, JPEG allowed.'}), 400

###############################################################################
#                  UPDATE STRUCTURED ANALYSIS ENDPOINT
###############################################################################
@app.route('/saveStructuredAnalysis', methods=['POST'])
@login_required
def save_structured_analysis():
    """Updates Firestore with new structured analysis."""
    try:
        username = session['username']
        data = request.get_json()
        thread_name = data.get('thread_name', '').strip()
        file_name = data.get('file_name', '').strip()
        updated_analysis = data.get('structured_analysis', '').strip()
        if not thread_name or not file_name or not updated_analysis:
            return jsonify({"error": "Missing required data."}), 400

        thread_ref = db.collection('users').document(username).collection('threads').document(thread_name)
        if not thread_ref.get().exists:
            return jsonify({'error': f'Thread "{thread_name}" not found'}), 404

        images_query = thread_ref.collection('images').where('file_name', '==', file_name).limit(1).stream()
        images_docs = list(images_query)
        if not images_docs:
            return jsonify({'error': f'No matching image with file_name "{file_name}" found.'}), 404

        image_doc_ref = images_docs[0].reference
        doc_data = images_docs[0].to_dict()

        # Find the last version used
        version_numbers = []
        for key in doc_data.keys():
            match = re.match(r'^structured_analysis_(\d+)$', key)
            if match:
                version_numbers.append(int(match.group(1)))

        # increment the version number
        new_version = max(version_numbers) + 1 if version_numbers else 1

        update_payload = {
            f"structured_analysis_{new_version}": updated_analysis,
            "latest_structured_analysis": updated_analysis,
            "updated_at": datetime.datetime.utcnow(),
            "structured_analysis": updated_analysis  # For backup reference
        }
        image_doc_ref.update(update_payload)

        logger.info(f"Structured analysis updated for file '{file_name}' in thread '{thread_name}'. Version = {new_version}")
        return jsonify({"message": "Structured analysis saved successfully."}), 200

    except Exception as e:
        logger.error(f"Error saving structured analysis: {str(e)}")
        return jsonify({'error': str(e)}), 500

###############################################################################
#                            DELETE IMAGE ENDPOINT
###############################################################################
@app.route('/deleteImage', methods=['POST'])
@login_required
def delete_image():
    """Deletes an image record and its file from Cloud Storage."""
    try:
        username = session['username']
        data = request.get_json()
        thread_name = data.get('thread_name', '').strip()
        file_name = data.get('file_name', '').strip()
        if not thread_name or not file_name:
            return jsonify({"error": "Missing required data."}), 400

        thread_ref = db.collection('users').document(username).collection('threads').document(thread_name)
        if not thread_ref.get().exists:
            return jsonify({'error': f'Thread "{thread_name}" not found'}), 404

        images_query = thread_ref.collection('images').where('file_name', '==', file_name).limit(1).stream()
        images_docs = list(images_query)
        if not images_docs:
            return jsonify({'error': f'No matching image with file_name "{file_name}" found.'}), 404

        image_doc_ref = images_docs[0].reference
        image_doc = images_docs[0].to_dict()
        gcs_path = image_doc.get('gcs_path', '')
        image_doc_ref.delete()

        if gcs_path and gcs_path.startswith('gs://'):
            parts = gcs_path.replace('gs://', '').split('/', 1)
            if len(parts) == 2:
                bucket_name, file_name_gcs = parts
                # Attempt to delete from storage
                storage_bucket = storage_client.bucket(bucket_name)
                blob = storage_bucket.blob(file_name_gcs)
                blob.delete()

        logger.info(f"Image '{file_name}' deleted from thread '{thread_name}' for user '{username}'.")
        return jsonify({'message': 'Image deleted successfully.'}), 200

    except Exception as e:
        logger.error(f"Error deleting image: {str(e)}")
        return jsonify({'error': str(e)}), 500

###############################################################################
#                   ADD ROWS TO STRUCTURED ANALYSIS JSON
###############################################################################
@app.route('/addNewRow', methods=['POST'])
@login_required
def add_new_row():
    """Handles adding a new row to the structured analysis JSON data."""
    try:
        username = session['username']
        data = request.get_json()
        thread_name = data.get('thread_name', '').strip()
        file_name = data.get('file_name', '').strip()
        new_row = data.get('new_row')

        if not thread_name or not file_name or not new_row:
            return jsonify({"error": "Missing required data."}), 400

        thread_ref = db.collection('users').document(username).collection('threads').document(thread_name)
        if not thread_ref.get().exists:
            return jsonify({'error': f'Thread "{thread_name}" not found'}), 404

        images_query = thread_ref.collection('images').where('file_name', '==', file_name).limit(1).stream()
        images_docs = list(images_query)
        if not images_docs:
            return jsonify({'error': f'No matching image with file_name "{file_name}" found.'}), 404

        image_doc_ref = images_docs[0].reference
        image_doc = images_docs[0].to_dict()

        latest_analysis = image_doc.get('latest_structured_analysis') or image_doc.get('structured_analysis')
        if not latest_analysis:
            return jsonify({"error": "No existing analysis found."}), 404

        try:
            analysis_json = json.loads(latest_analysis)
        except json.JSONDecodeError:
            return jsonify({"error": "Invalid JSON format."}), 400

        rows = analysis_json.get('rows', [])

        # Figure out the next serial number
        new_serial_number = max((r.get('serial_number', 0) for r in rows), default=0) + 1
        new_row['serial_number'] = new_serial_number

        rows.append(new_row)
        analysis_json['rows'] = rows
        updated_analysis = json.dumps(analysis_json)

        update_payload = {
            "latest_structured_analysis": updated_analysis,
            "updated_at": datetime.datetime.utcnow(),
            "structured_analysis": updated_analysis
        }
        image_doc_ref.update(update_payload)

        logger.info(f"New row added to analysis for file '{file_name}' in thread '{thread_name}'.")
        return jsonify({"message": "New row added successfully."}), 200

    except Exception as e:
        logger.error(f"Error adding new row: {str(e)}")
        return jsonify({'error': str(e)}), 500

@ app.route('/downloadAnalysis', methods=['GET'])
@ login_required
def download_analysis():
    """Generates an Excel file with combined structured analysis of all images in a thread."""
    try:
        username = session['username']
        thread_name = request.args.get('threadName')
        if not thread_name:
            return jsonify({'error': 'threadName parameter is required'}), 400

        thread_ref = db.collection('users').document(username).collection('threads').document(thread_name)
        if not thread_ref.get().exists:
            return jsonify({'error': f'Thread "{thread_name}" not found'}), 404

        images_collection = thread_ref.collection('images').stream()
        per_image_data = []  # List to hold data for each image

        for img_doc in images_collection:
            img_data = img_doc.to_dict()
            latest_json = img_data.get('latest_structured_analysis') or img_data.get('structured_analysis')
            if not latest_json:
                continue
            try:
                parsed_json = json.loads(latest_json)
            except json.JSONDecodeError:
                logger.error("Invalid JSON in structured_analysis or latest_structured_analysis.")
                continue

            rows = parsed_json.get('rows', [])
            # Collect ONLY "name" and "quantity" for the Excel rows
            excel_data = []
            for row in rows:
                excel_data.append({
                    'Name': row.get("name", ""),
                    'Quantity': row.get("quantity", "")
                })

            additional_notes = parsed_json.get('additional_notes', [])

            per_image_data.append({
                'file_name': img_data.get('file_name', 'Unnamed Image'),
                'excel_data': excel_data,
                'additional_notes': additional_notes
            })

        if not per_image_data:
            return jsonify({'error': 'No valid structured analysis found for images in this thread.'}), 404

        excel_file = json_to_excel(per_image_data)
        return excel_file

    except Exception as e:
        logger.error(f"Error generating download analysis: {str(e)}")
        return jsonify({'error': str(e)}), 500


def json_to_excel(per_image_data):
    """Converts the JSON-based structured analysis to an Excel file (openpyxl),
       containing component and quantity data followed by additional notes for each image,
       separated by a yellow line.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define headers
    headers = ["Name", "Quantity"]
    ws.append(headers)

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True, size=12)

    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Start writing data from the second row
    current_row = 2

    for image_data in per_image_data:
        file_name = image_data.get('file_name', 'Unnamed Image')
        excel_rows = image_data.get('excel_data', [])
        additional_notes = image_data.get('additional_notes', [])

        # Add Image Title
        ws.cell(row=current_row, column=1, value=f"Component and Quantity Data - {file_name}")
        ws.cell(row=current_row, column=1).font = bold_font
        current_row += 1

        # Add component and quantity data
        for row_data in excel_rows:
            ws.cell(row=current_row, column=1, value=row_data.get("Name", ""))
            ws.cell(row=current_row, column=2, value=row_data.get("Quantity", ""))
            current_row += 1

        # Add Additional Notes Header
        ws.cell(row=current_row, column=1, value="Additional Notes:")
        ws.cell(row=current_row, column=1).font = bold_font
        current_row += 1

        # Add Additional Notes
        for note in additional_notes:
            ws.cell(row=current_row, column=1, value=note)
            current_row += 1

        # Add Yellow Separator Row
        for col in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=col).fill = yellow_fill
        current_row += 1

    # Apply borders and auto-width
    for row in ws.iter_rows(min_row=1, max_row=current_row - 1):
        for cell in row:
            cell.border = thin_border

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = length + 2
        column_letter = openpyxl.utils.get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Write to BytesIO
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    # Prepare response
    response = make_response(excel_stream.read())
    response.headers.set('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response.headers.set('Content-Disposition', 'attachment; filename=analysis.xlsx')
    return response


###############################################################################
#                               FLASK MAIN
###############################################################################
if __name__ == '__main__':
    app.run(debug=True)